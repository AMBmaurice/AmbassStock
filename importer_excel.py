import os
import django
import openpyxl

# Configuration de l'environnement Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from config.models import Produit

def importer(chemin_fichier):
    wb = openpyxl.load_workbook(chemin_fichier)
    
    # On force le script à ouvrir spécifiquement la feuille "Inventaire "
    if "Inventaire " in wb.sheetnames:
        sheet = wb["Inventaire "]
    else:
        sheet = wb.active
        
    print("Début de l'importation depuis la feuille Inventaire...")
    
    # On commence à la ligne 3 (car ligne 1 = vide, ligne 2 = titres)
    for row in sheet.iter_rows(min_row=3, values_only=True):
        # Si la ligne entière ou la référence est vide, on passe à la suivante
        if not row or row[0] is None:
            continue
            
        ref = str(row[0]).strip()
        obj = str(row[1]).strip()
        
        # Sécurité pour la quantité actuelle (colonne 4 dans ton Excel)
        try:
            qte = int(row[5]) if row[5] is not None else 0
        except (ValueError, TypeError):
            qte = 0
            
        cat = str(row[3]).strip() if row[3] else "Divers"
        emp = str(row[4]).strip() if row[4] else "Non spécifié"
        
        # Enregistrement ou mise à jour sécurisée dans Django
        Produit.objects.update_or_create(
            reference=ref,
            defaults={
                'objet': obj,
                'quantite': qte,
                'categorie': cat,
                'emplacement': emp
            }
        )
        print(f"✅ Importé : {obj} [{ref}] - Stock : {qte} ({emp})")

    print("\n🎉 Félicitations ! Ton vrai catalogue d'ambassade est chargé dans Django.")

if __name__ == '__main__':
    importer('inventaire.xlsx')
